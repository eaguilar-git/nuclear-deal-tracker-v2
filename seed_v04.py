"""
seed_v04.py
═══════════════════════════════════════════════════════════════════════════════
One-time script to populate a fresh Google Sheet with the v0.4 6-tab schema
seeded from the Pass 5 Excel database.

This is the foundation step for the Nuclear Deal Tracker v2 system. Run once
against a brand-new Google Sheet to bootstrap it with:
  - 69 Sites
  - 126 Reactor Units
  - 47 Projects
  - 35 Deals (all marked confirmation_status="Confirmed")
  - 48 Context Items
  - 114 Announcements
  Plus operational tabs: Seen (deduplication), Review (queue)

Schema differences from Pass 5 Excel:
  - Deals tab gets new `confirmation_status` column (values: Confirmed | Proposed (scraper))
    Seed deals are all "Confirmed". Scraper-proposed deals will be tagged "Proposed (scraper)"
    for human review before promotion.
  - Announcements tab drops `raw_body` column (was empty for backfilled rows;
    keeps Sheet size manageable for daily scraping).

Usage
─────
  pip install openpyxl gspread google-auth

  export GOOGLE_SERVICE_ACCOUNT_JSON=$(cat your-service-account-key.json)
  export GOOGLE_SHEET_ID=your_new_sheet_id

  python seed_v04.py

  Optional flags:
    --excel <path>        Path to Pass 5 Excel (default: ./reference/Nuclear_Deal_Tracker_v2_Pass5.xlsx)
    --dry-run             Print what would be written, don't touch the Sheet
    --confirm-overwrite   Required if any target tab already has data

Setup notes
───────────
1. Create a fresh Google Sheet (blank).
2. Create a service account in Google Cloud Console, enable Sheets API,
   download the JSON key.
3. Share the Sheet with the service account email (Editor access).
4. Set the env vars above and run.

Author: Edgar Aguilar / EFI Foundation
Schema reference: Nuclear_Deal_Tracker_Schema_Backbone_v0.4_locked.docx
"""

import os
import sys
import json
import time
import argparse
from pathlib import Path

import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook


# ─── CONFIG ──────────────────────────────────────────────────────────────────

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

# Tab order (left to right in the Sheet)
TAB_ORDER = [
    "Sites",
    "Reactor Units",
    "Projects",
    "Deals",
    "Context Items",
    "Announcements",
    "Seen",
    "Review",
]

# Schema deltas vs Pass 5 Excel
DROP_COLUMNS = {
    "Announcements": ["raw_body"],
}

# New columns to ADD (tab → list of (column_name, default_value, insert_after))
ADD_COLUMNS = {
    "Deals": [
        ("confirmation_status", "Confirmed", "deal_id"),  # right after deal_id
    ],
}

# Operational tabs (created fresh, not seeded from Excel)
OPERATIONAL_TABS = {
    "Seen": ["hash", "title", "url", "scraped_at"],
    "Review": [
        "review_id",
        "scraped_at",
        "article_title",
        "article_url",
        "reason",
        "raw_extraction",
    ],
}


# ─── EXCEL LOADING ──────────────────────────────────────────────────────────

def load_excel_tab(wb, tab_name):
    """Returns (headers, rows) where rows is a list of lists matching headers."""
    ws = wb[tab_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = list(rows[0])
    data = []
    for row in rows[1:]:
        # Skip fully empty rows
        if not row or row[0] is None or row[0] == "":
            continue
        # Convert None → ""
        clean_row = ["" if v is None else v for v in row]
        # Pad/truncate to header length
        if len(clean_row) < len(headers):
            clean_row += [""] * (len(headers) - len(clean_row))
        elif len(clean_row) > len(headers):
            clean_row = clean_row[:len(headers)]
        data.append(clean_row)
    return headers, data


def apply_schema_deltas(tab_name, headers, rows):
    """Apply DROP_COLUMNS and ADD_COLUMNS to a loaded tab."""
    headers = list(headers)
    rows = [list(r) for r in rows]

    # 1) Drop columns
    drops = DROP_COLUMNS.get(tab_name, [])
    for col_name in drops:
        if col_name in headers:
            idx = headers.index(col_name)
            headers.pop(idx)
            for r in rows:
                if idx < len(r):
                    r.pop(idx)

    # 2) Add columns
    adds = ADD_COLUMNS.get(tab_name, [])
    for col_name, default_val, after_col in adds:
        if col_name in headers:
            continue  # already there
        if after_col not in headers:
            # Append at end if anchor not found
            insert_at = len(headers)
        else:
            insert_at = headers.index(after_col) + 1
        headers.insert(insert_at, col_name)
        for r in rows:
            r.insert(insert_at, default_val)

    return headers, rows


# ─── GOOGLE SHEETS ──────────────────────────────────────────────────────────

def connect_sheet(sheet_id):
    sa_json = os.environ["GOOGLE_SERVICE_ACCOUNT_JSON"]
    creds = Credentials.from_service_account_info(
        json.loads(sa_json),
        scopes=SCOPES,
    )
    gc = gspread.authorize(creds)
    return gc.open_by_key(sheet_id)


def get_or_create_tab(spreadsheet, tab_name, expected_cols):
    """Return worksheet handle. Create if missing. Return existing if present."""
    existing = [ws.title for ws in spreadsheet.worksheets()]
    if tab_name in existing:
        return spreadsheet.worksheet(tab_name), False
    ws = spreadsheet.add_worksheet(
        title=tab_name,
        rows=2000,
        cols=max(expected_cols + 4, 20),
    )
    return ws, True


def has_data(ws):
    """Check if a worksheet has any data rows beyond header."""
    try:
        # Read first 2 rows; if row 2 has any content, there's data
        vals = ws.get("A2:Z2")
        if not vals:
            return False
        return any(cell.strip() for cell in vals[0]) if vals else False
    except Exception:
        return False


def write_tab(ws, headers, rows, tab_name):
    """Clear sheet and write headers + rows."""
    ws.clear()
    time.sleep(0.5)

    # Write header
    ws.update("A1", [headers], value_input_option="RAW")
    time.sleep(0.5)

    # Write data in chunks (gspread can handle ~5k cells per call)
    if rows:
        # Convert all values to strings for predictable behavior
        # Numbers will be re-interpreted by Sheets but the strings get stored cleanly
        str_rows = [
            [str(v) if v not in (None, "") else "" for v in r]
            for r in rows
        ]
        # Determine end column letter
        end_col = chr(ord("A") + len(headers) - 1) if len(headers) <= 26 else "AZ"
        end_row = 1 + len(str_rows)
        if len(headers) > 26:
            # Multi-letter column logic for >26 cols
            n = len(headers) - 1
            first = chr(ord("A") + n // 26 - 1) if n >= 26 else ""
            second = chr(ord("A") + n % 26)
            end_col = f"{first}{second}"

        # Use append_rows in chunks if very large
        if len(str_rows) <= 200:
            ws.update(f"A2:{end_col}{end_row}", str_rows, value_input_option="USER_ENTERED")
        else:
            # Chunk to avoid timeouts
            chunk = 100
            for i in range(0, len(str_rows), chunk):
                batch = str_rows[i:i + chunk]
                start = i + 2
                end = start + len(batch) - 1
                ws.update(
                    f"A{start}:{end_col}{end}",
                    batch,
                    value_input_option="USER_ENTERED",
                )
                time.sleep(1.5)  # rate-limit cushion

        time.sleep(1)


def reorder_tabs(spreadsheet, desired_order):
    """Reorder worksheets to match desired_order. Tabs not in the list go to the end."""
    existing = {ws.title: ws for ws in spreadsheet.worksheets()}
    for i, tab_name in enumerate(desired_order):
        ws = existing.get(tab_name)
        if ws is None:
            continue
        try:
            spreadsheet.reorder_worksheets([
                existing[t] for t in desired_order if t in existing
            ] + [ws for t, ws in existing.items() if t not in desired_order])
            return  # reorder_worksheets does it in one shot
        except Exception:
            pass


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Seed a fresh Google Sheet with v0.4 schema")
    parser.add_argument(
        "--excel",
        default="./reference/Nuclear_Deal_Tracker_v2_Pass5.xlsx",
        help="Path to Pass 5 Excel file",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print what would be written, don't touch the Sheet",
    )
    parser.add_argument(
        "--confirm-overwrite",
        action="store_true",
        help="Required if any target tab already has data",
    )
    args = parser.parse_args()

    # Validate inputs
    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"❌ Excel file not found: {excel_path}")
        print(f"   Tip: pass --excel <path> to point at the correct file")
        sys.exit(1)

    if not args.dry_run:
        for var in ("GOOGLE_SERVICE_ACCOUNT_JSON", "GOOGLE_SHEET_ID"):
            if var not in os.environ:
                print(f"❌ Environment variable not set: {var}")
                sys.exit(1)

    print("═" * 70)
    print("Nuclear Deal Tracker v2 — Sheet seeder (v0.4 schema)")
    print("═" * 70)
    print(f"  Source: {excel_path}")
    if args.dry_run:
        print(f"  Mode:   DRY RUN (no writes)")
    else:
        sheet_id = os.environ["GOOGLE_SHEET_ID"]
        print(f"  Target: Google Sheet {sheet_id}")
    print()

    # ─── Load Excel ──────────────────────────────────────────────────────
    print("Loading Pass 5 Excel...")
    wb = load_workbook(excel_path, read_only=True, data_only=True)

    seed_data = {}
    for tab in ["Sites", "Reactor Units", "Projects", "Deals", "Context Items", "Announcements"]:
        if tab not in wb.sheetnames:
            print(f"  ❌ Tab '{tab}' not found in Excel")
            sys.exit(1)
        headers, rows = load_excel_tab(wb, tab)
        headers, rows = apply_schema_deltas(tab, headers, rows)
        seed_data[tab] = (headers, rows)
        delta_note = ""
        drops = DROP_COLUMNS.get(tab, [])
        adds = [a[0] for a in ADD_COLUMNS.get(tab, [])]
        if drops or adds:
            parts = []
            if drops:
                parts.append(f"dropped: {', '.join(drops)}")
            if adds:
                parts.append(f"added: {', '.join(adds)}")
            delta_note = f"  [{'; '.join(parts)}]"
        print(f"  ✓ {tab}: {len(rows)} rows, {len(headers)} columns{delta_note}")

    # ─── Validate FK integrity in Excel before writing ───────────────────
    print("\nValidating FK integrity (pre-flight check)...")

    valid_sites = set(r[0] for r in seed_data["Sites"][1] if r[0])
    valid_units = set(r[0] for r in seed_data["Reactor Units"][1] if r[0])
    valid_projects = set(r[0] for r in seed_data["Projects"][1] if r[0])
    valid_deals = set(r[0] for r in seed_data["Deals"][1] if r[0])
    valid_contexts = set(r[0] for r in seed_data["Context Items"][1] if r[0])

    fk_errors = []

    # Reactor Units → Sites
    ru_headers, ru_rows = seed_data["Reactor Units"]
    site_id_idx = ru_headers.index("site_id")
    for r in ru_rows:
        if r[site_id_idx] and r[site_id_idx] not in valid_sites:
            fk_errors.append(f"Reactor Unit '{r[0]}' → site '{r[site_id_idx]}' not found")

    # Projects → Reactor Units (linked_unit_ids is comma-separated)
    p_headers, p_rows = seed_data["Projects"]
    luid_idx = p_headers.index("linked_unit_ids")
    for r in p_rows:
        if r[luid_idx]:
            for uid in str(r[luid_idx]).split(","):
                uid = uid.strip()
                if uid and uid not in valid_units:
                    fk_errors.append(f"Project '{r[0]}' → unit '{uid}' not found")

    # Deals → Projects + Context Items
    d_headers, d_rows = seed_data["Deals"]
    pids_idx = d_headers.index("project_ids")
    cids_idx = d_headers.index("context_ids")
    for r in d_rows:
        if r[pids_idx]:
            for pid in str(r[pids_idx]).split(","):
                pid = pid.strip()
                if pid and pid not in valid_projects:
                    fk_errors.append(f"Deal '{r[0]}' → project '{pid}' not found")
        if r[cids_idx]:
            for cid in str(r[cids_idx]).split(","):
                cid = cid.strip()
                if cid and cid not in valid_contexts:
                    fk_errors.append(f"Deal '{r[0]}' → context '{cid}' not found")

    # Announcements → all four
    a_headers, a_rows = seed_data["Announcements"]
    for col_name, valid_set, label in [
        ("project_ids", valid_projects, "project"),
        ("context_ids", valid_contexts, "context"),
        ("deal_ids", valid_deals, "deal"),
        ("unit_ids", valid_units, "unit"),
        ("site_ids", valid_sites, "site"),
    ]:
        col_idx = a_headers.index(col_name)
        for r in a_rows:
            if r[col_idx]:
                for ref in str(r[col_idx]).split(","):
                    ref = ref.strip()
                    if ref and ref not in valid_set:
                        fk_errors.append(f"Announcement '{r[0]}' → {label} '{ref}' not found")

    if fk_errors:
        print(f"  ❌ {len(fk_errors)} FK errors found:")
        for err in fk_errors[:10]:
            print(f"     - {err}")
        if len(fk_errors) > 10:
            print(f"     ... and {len(fk_errors) - 10} more")
        print()
        print("  Aborting. Fix the Excel source before seeding.")
        sys.exit(1)
    print("  ✓ All FK references valid")

    # ─── Dry run exit ────────────────────────────────────────────────────
    if args.dry_run:
        print("\n" + "═" * 70)
        print("DRY RUN — what would be written:")
        print("═" * 70)
        for tab in TAB_ORDER:
            if tab in seed_data:
                headers, rows = seed_data[tab]
                print(f"  {tab:18s}  {len(rows):>4} rows × {len(headers)} cols")
            elif tab in OPERATIONAL_TABS:
                print(f"  {tab:18s}     (empty operational tab — headers only)")
        print("\nRun without --dry-run to actually seed the Sheet.")
        return

    # ─── Connect to Sheet ────────────────────────────────────────────────
    print("\nConnecting to Google Sheet...")
    spreadsheet = connect_sheet(os.environ["GOOGLE_SHEET_ID"])
    print(f"  ✓ Connected: {spreadsheet.title}")

    # ─── Pre-flight: check for existing data ─────────────────────────────
    print("\nChecking existing tabs...")
    existing_tabs = {ws.title: ws for ws in spreadsheet.worksheets()}
    tabs_with_data = []
    for tab in TAB_ORDER:
        if tab in existing_tabs:
            ws = existing_tabs[tab]
            if has_data(ws):
                tabs_with_data.append(tab)

    if tabs_with_data:
        print(f"  ⚠ Found existing data in: {', '.join(tabs_with_data)}")
        if not args.confirm_overwrite:
            print(
                "  Refusing to overwrite without --confirm-overwrite flag.\n"
                "  Re-run with --confirm-overwrite to proceed."
            )
            sys.exit(1)
        print("  ✓ --confirm-overwrite set; will replace existing data")
    else:
        print("  ✓ No existing data; safe to seed")

    # ─── Seed data tabs ──────────────────────────────────────────────────
    print("\nSeeding data tabs...")
    for tab in ["Sites", "Reactor Units", "Projects", "Deals", "Context Items", "Announcements"]:
        headers, rows = seed_data[tab]
        ws, created = get_or_create_tab(spreadsheet, tab, len(headers))
        action = "Created and seeding" if created else "Overwriting"
        print(f"  {action} '{tab}' ({len(rows)} rows)...")
        write_tab(ws, headers, rows, tab)
        print(f"  ✓ '{tab}' done")

    # ─── Create operational tabs ─────────────────────────────────────────
    print("\nCreating operational tabs...")
    for tab, headers in OPERATIONAL_TABS.items():
        ws, created = get_or_create_tab(spreadsheet, tab, len(headers))
        if created or has_data(ws):
            ws.clear()
            ws.update("A1", [headers], value_input_option="RAW")
            print(f"  ✓ '{tab}' (headers only)")
        else:
            ws.update("A1", [headers], value_input_option="RAW")
            print(f"  ✓ '{tab}' headers refreshed")
        time.sleep(0.5)

    # ─── Reorder tabs ────────────────────────────────────────────────────
    try:
        existing = {ws.title: ws for ws in spreadsheet.worksheets()}
        ordered = [existing[t] for t in TAB_ORDER if t in existing]
        # Append any extras at the end
        ordered += [ws for t, ws in existing.items() if t not in TAB_ORDER]
        spreadsheet.reorder_worksheets(ordered)
        print("\n✓ Tabs reordered.")
    except Exception as e:
        print(f"\n⚠ Could not reorder tabs ({e}). Drag manually if order matters.")

    print("\n" + "═" * 70)
    print(f"✅ Seed complete.")
    print(f"   Sheet URL: https://docs.google.com/spreadsheets/d/{os.environ['GOOGLE_SHEET_ID']}/edit")
    print("═" * 70)
    print("\nNext step: publish each data tab to web as CSV (File → Share → Publish to web)")
    print("           and capture the resulting gids for the dashboard's data fetch URLs.")


if __name__ == "__main__":
    main()
